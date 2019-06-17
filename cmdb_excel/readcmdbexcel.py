import openpyxl

workbook = openpyxl.load_workbook("file/20190408.xlsx")

sheet_names = workbook.get_sheet_names()

work_sheet = workbook.get_sheet_by_name("详表")

rows = work_sheet.max_row

columns = work_sheet.max_column

for row in work_sheet.rows:
    for cell in row:
        if cell.value != None:
            print(cell, cell.value, end=" ")
    print()

print(sheet_names)
print(work_sheet)
print(rows, columns)

